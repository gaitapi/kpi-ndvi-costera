import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_00
import numpy as np

# ── Datos ─────────────────────────────────────────────────────────────────────
df = pd.read_csv(
    '/home/gaitapi/proyectos/kpi-costa/script para correr en GEE el NDVI de todos los tramos costeros 45 metros/KPI_NDVI_tendencia.csv'
)
# Normalizar nombre de departamento
df['d'] = df['d'].str.strip().str.lower().str.replace('san josé','san jose')

# ── Paleta ────────────────────────────────────────────────────────────────────
C = {
    'header_bg':    '0D2640',
    'header_fg':    'EEE8DC',
    'subhdr_ndvi':  '144A7A',
    'subhdr_tend':  '1A5C3A',
    'subhdr_inst':  '6B4C1E',
    'subhdr_rest':  '4A1E6B',
    'alt_row':      'F5F8FC',
    'white':        'FFFFFF',
    'vul1_bg':      'FCE4EC',  # Alta — rojo suave
    'vul2_bg':      'FFF9C4',  # Media — amarillo
    'vul3_bg':      'E8F5E9',  # Baja — verde
    'tend_mej':     'C8E6C9',
    'tend_est':     'FFF9C4',
    'tend_deg':     'FFCDD2',
    'border':       'BFCAD8',
    'ndvi_hi':      '2E7D32',  # verde oscuro (NDVI alto)
    'ndvi_lo':      'B71C1C',  # rojo oscuro (NDVI bajo)
}

DEPTS = ['colonia','san jose','montevideo','canelones','maldonado','rocha']
DEPT_NAMES = {
    'colonia':    'Colonia',
    'san jose':   'San José',
    'montevideo': 'Montevideo',
    'canelones':  'Canelones',
    'maldonado':  'Maldonado',
    'rocha':      'Rocha',
}
VUL_LABEL = {1:'Alta', 2:'Media', 3:'Baja'}
ACC_LABEL = {0:'Sin acceso', 1:'Bajo', 2:'Medio', 3:'Alto'}

NDVI_YEARS = [2017,2018,2019,2020,2021,2022,2023,2024]

INSTRUMENTOS = [
    'Plan de manejo costero',
    'Guía / protocolo de manejo',
    'Lineamientos de gestión',
    'Anteproyecto de obra',
    'Proyecto ejecutivo',
]
INST_OPTS = 'No existe,En elaboración,Vigente,Desactualizado'

ACCIONES = [
    'Plantación / Revegetación',
    'Retiro de especies exóticas',
    'Manejo de accesos',
    'Manejo de pisoteo / senderos',
    'Obra de estabilización',
    'Monitoreo activo',
    'Otras acciones',
]
ACCION_OPTS = 'No,Sí'

# ── Helpers ───────────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def font(bold=False, color='000000', size=10, name='Arial'):
    return Font(bold=bold, color=color, size=size, name=name)

def align(h='center', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(color='BFCAD8'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def ndvi_color(val):
    """Interpolar color entre rojo (0) y verde (0.8+)."""
    if pd.isna(val): return 'D0D0D0'
    t = min(max(val / 0.75, 0), 1)
    r = int(183 * (1-t) + 46 * t)
    g = int(28  * (1-t) + 125 * t)
    b = int(28  * (1-t) + 50  * t)
    return f'{r:02X}{g:02X}{b:02X}'

def set_header_cell(ws, row, col, text, bg, fg='EEE8DC', bold=True,
                    size=9, wrap=True, h='center'):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(bg)
    c.font = font(bold=bold, color=fg, size=size)
    c.alignment = align(h=h, v='center', wrap=wrap)
    c.border = thin_border()
    return c

# ── Construcción ──────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)  # quitar hoja vacía inicial

for dept_key in DEPTS:
    dept_df = df[df['d'] == dept_key].copy().reset_index(drop=True)
    if dept_df.empty:
        continue

    ws = wb.create_sheet(title=DEPT_NAMES[dept_key])
    ws.freeze_panes = 'A4'  # congelar filas de cabecera

    # ── Definir columnas ──────────────────────────────────────────────────────
    # Fila 1: grupos principales
    # Fila 2: subgrupos
    # Fila 3: nombres de columna
    # Fila 4+: datos

    # Índices de columna (1-based)
    COL_TRAMO    = 1   # t
    COL_LARGO    = 2   # l
    COL_VUL      = 3   # vulnerabilidad
    COL_ACC      = 4   # accesibilidad
    ndvi_start   = 5
    ndvi_end     = ndvi_start + len(NDVI_YEARS) - 1   # 5..12
    COL_TEND     = ndvi_end + 1   # 13
    COL_SLOPE    = ndvi_end + 2   # 14
    COL_MKP      = ndvi_end + 3   # 15
    inst_start   = COL_MKP + 1    # 16
    inst_end     = inst_start + len(INSTRUMENTOS) - 1  # 16..20
    rest_start   = inst_end + 1   # 21
    # cada acción ocupa 2 columnas: Realizada + Año
    rest_end     = rest_start + len(ACCIONES)*2 - 1
    # "otras" tiene 3 cols: Realizada + Año + Descripción
    # hacemos la última acción (Otras) con columna extra
    rest_end     += 1  # +1 para descripción de "Otras"
    TOTAL_COLS   = rest_end

    # ── Fila 1: Grupos ────────────────────────────────────────────────────────
    grp = [
        (COL_TRAMO, COL_ACC,    'IDENTIFICACIÓN',  C['header_bg']),
        (ndvi_start, ndvi_end,  'NDVI ANUAL · Sentinel-2 · Buffer 45 m', C['subhdr_ndvi']),
        (COL_TEND, COL_MKP,     'TENDENCIA',       C['subhdr_tend']),
        (inst_start, inst_end,  'INSTRUMENTOS DE MANEJO', C['subhdr_inst']),
        (rest_start, rest_end,  'ACCIONES DE RESTAURACIÓN', C['subhdr_rest']),
    ]
    for c1, c2, label, bg in grp:
        ws.merge_cells(start_row=1, start_column=c1,
                       end_row=1,   end_column=c2)
        set_header_cell(ws, 1, c1, label, bg, size=10, bold=True)

    # ── Fila 2: Subgrupos para acciones (Acción / Año) ────────────────────────
    # Celdas vacías del mismo color para filas 1-2 en otras secciones
    for col in range(COL_TRAMO, TOTAL_COLS+1):
        c = ws.cell(row=2, column=col)
        c.border = thin_border()
        c.alignment = align()
    # Subgrupos de instrumentos (fila 2 = etiqueta de subgrupo)
    for i, inst in enumerate(INSTRUMENTOS):
        col = inst_start + i
        set_header_cell(ws, 2, col, inst, C['subhdr_inst'], size=8, wrap=True)
    # Subgrupos de acciones
    for i, acc in enumerate(ACCIONES):
        c1 = rest_start + i*2
        ws.merge_cells(start_row=2, start_column=c1,
                       end_row=2,   end_column=c1+1)
        set_header_cell(ws, 2, c1, acc, C['subhdr_rest'], size=8, wrap=True)
    # Última acción (Otras) tiene descripción extra → 3 cols
    other_start = rest_start + (len(ACCIONES)-1)*2
    ws.merge_cells(start_row=2, start_column=other_start,
                   end_row=2,   end_column=other_start+2)
    set_header_cell(ws, 2, other_start, 'Otras acciones', C['subhdr_rest'], size=8)

    # Rellenar filas 1-2 para columnas base e NDVI
    for col in range(COL_TRAMO, ndvi_end+1):
        for row in [1,2]:
            c = ws.cell(row=row, column=col)
            if not c.value:
                bg = C['header_bg'] if col <= COL_ACC else C['subhdr_ndvi']
                c.fill = fill(bg)
                c.border = thin_border()
    for col in range(COL_TEND, COL_MKP+1):
        for row in [1,2]:
            c = ws.cell(row=row, column=col)
            if not c.value:
                c.fill = fill(C['subhdr_tend'])
                c.border = thin_border()

    # ── Fila 3: Nombres de columna ────────────────────────────────────────────
    headers_row3 = {
        COL_TRAMO: 'Tramo',
        COL_LARGO: 'Largo (m)',
        COL_VUL:   'Vulnerabilidad',
        COL_ACC:   'Accesibilidad',
        COL_TEND:  'Tendencia',
        COL_SLOPE: 'Slope Sen',
        COL_MKP:   'MK p-val',
    }
    for col, label in headers_row3.items():
        bg = C['header_bg'] if col <= COL_ACC else C['subhdr_tend']
        if col >= ndvi_start and col <= ndvi_end:
            bg = C['subhdr_ndvi']
        set_header_cell(ws, 3, col, label, bg, size=9)
    for i, yr in enumerate(NDVI_YEARS):
        set_header_cell(ws, 3, ndvi_start+i, str(yr), C['subhdr_ndvi'], size=9)
    for i in range(len(INSTRUMENTOS)):
        set_header_cell(ws, 3, inst_start+i, 'Estado', C['subhdr_inst'], size=8)
    for i, acc in enumerate(ACCIONES[:-1]):
        base = rest_start + i*2
        set_header_cell(ws, 3, base,   'Realizada', C['subhdr_rest'], size=8)
        set_header_cell(ws, 3, base+1, 'Año',       C['subhdr_rest'], size=8)
    other_base = rest_start + (len(ACCIONES)-1)*2
    set_header_cell(ws, 3, other_base,   'Realizada',   C['subhdr_rest'], size=8)
    set_header_cell(ws, 3, other_base+1, 'Año',         C['subhdr_rest'], size=8)
    set_header_cell(ws, 3, other_base+2, 'Descripción', C['subhdr_rest'], size=8)

    # ── Validaciones (DataValidation) ─────────────────────────────────────────
    # Instrumentos
    dv_inst = DataValidation(
        type='list', formula1=f'"{INST_OPTS}"',
        allow_blank=True, showErrorMessage=False
    )
    ws.add_data_validation(dv_inst)
    # Acciones (Sí/No)
    dv_acc = DataValidation(
        type='list', formula1=f'"{ACCION_OPTS}"',
        allow_blank=True, showErrorMessage=False
    )
    ws.add_data_validation(dv_acc)
    # Año (número entre 2000 y 2035)
    dv_year = DataValidation(
        type='whole', operator='between',
        formula1='2000', formula2='2035',
        allow_blank=True, showErrorMessage=True,
        errorTitle='Año inválido', error='Ingresá un año entre 2000 y 2035'
    )
    ws.add_data_validation(dv_year)

    # ── Datos ─────────────────────────────────────────────────────────────────
    for ridx, row_data in dept_df.iterrows():
        excel_row = ridx + 4  # filas 1-3 = cabeceras
        is_alt = ridx % 2 == 1

        vul_val = int(row_data['v']) if not pd.isna(row_data['v']) else 1
        vul_bg = {'Alta':'vul1_bg','Media':'vul2_bg','Baja':'vul3_bg'}
        row_bg = C[f'vul{vul_val}_bg'] if not is_alt else C['alt_row']

        def data_cell(col, value, fmt=None, bold=False, h='center'):
            c = ws.cell(row=excel_row, column=col, value=value)
            c.fill = fill(row_bg)
            c.font = font(bold=bold, color='1A2636', size=9)
            c.alignment = align(h=h, v='center')
            c.border = thin_border()
            if fmt: c.number_format = fmt
            return c

        # Base
        data_cell(COL_TRAMO, int(row_data['t']), bold=True)
        data_cell(COL_LARGO, int(row_data['l']), fmt='#,##0')
        # Vulnerabilidad con color
        vul_c = ws.cell(row=excel_row, column=COL_VUL,
                        value=VUL_LABEL.get(vul_val, str(vul_val)))
        vul_bg_colors = {1: C['vul1_bg'], 2: C['vul2_bg'], 3: C['vul3_bg']}
        vul_fg_colors = {1: '7B0000', 2: '5D4000', 3: '1B5E20'}
        vul_c.fill = fill(vul_bg_colors.get(vul_val, C['white']))
        vul_c.font = font(bold=True, color=vul_fg_colors.get(vul_val,'000000'), size=9)
        vul_c.alignment = align()
        vul_c.border = thin_border()

        acc_val = int(row_data['a']) if not pd.isna(row_data['a']) else 0
        data_cell(COL_ACC, ACC_LABEL.get(acc_val, str(acc_val)))

        # NDVI por año con color de celda
        for yi, yr in enumerate(NDVI_YEARS):
            col = ndvi_start + yi
            col_name = f'NDVI_{yr}'
            val = row_data.get(col_name, None)
            if pd.notna(val):
                c = ws.cell(row=excel_row, column=col, value=round(float(val),4))
                ndvi_bg = ndvi_color(float(val))
                # Calcular si texto debe ser claro u oscuro
                r_v = int(ndvi_bg[0:2],16)
                g_v = int(ndvi_bg[2:4],16)
                lum = 0.299*r_v + 0.587*g_v + 0.114*int(ndvi_bg[4:6],16)
                fg_ndvi = 'FFFFFF' if lum < 128 else '1A2636'
                c.fill = fill(ndvi_bg)
                c.font = font(color=fg_ndvi, size=8)
                c.alignment = align()
                c.border = thin_border()
                c.number_format = '0.0000'
            else:
                data_cell(col, None)

        # Tendencia
        tend_val = str(row_data.get('tendencia','')).lower()
        tend_bg = {'mejora': C['tend_mej'], 'estable': C['tend_est'],
                   'degradacion': C['tend_deg']}.get(tend_val, C['white'])
        tend_fg = {'mejora':'1B5E20','estable':'5D4000','degradacion':'7B0000'}.get(tend_val,'000000')
        t_c = ws.cell(row=excel_row, column=COL_TEND, value=tend_val.capitalize())
        t_c.fill = fill(tend_bg)
        t_c.font = font(bold=True, color=tend_fg, size=9)
        t_c.alignment = align()
        t_c.border = thin_border()

        slope_val = row_data.get('sens_slope', None)
        data_cell(COL_SLOPE, round(float(slope_val),6) if pd.notna(slope_val) else None,
                  fmt='0.000000')
        mkp_val = row_data.get('mk_p', None)
        data_cell(COL_MKP, round(float(mkp_val),4) if pd.notna(mkp_val) else None,
                  fmt='0.0000')

        # Instrumentos (dropdown, celdas vacías editables)
        for i in range(len(INSTRUMENTOS)):
            col = inst_start + i
            c = ws.cell(row=excel_row, column=col, value='No existe')
            c.fill = fill(C['alt_row'] if is_alt else C['white'])
            c.font = font(color='5A3E00', size=9)
            c.alignment = align()
            c.border = thin_border()
            dv_inst.add(c)

        # Acciones (dropdown Sí/No + año, celdas vacías)
        for i, acc_name in enumerate(ACCIONES[:-1]):
            base = rest_start + i*2
            # Realizada
            cr = ws.cell(row=excel_row, column=base, value='No')
            cr.fill = fill(C['alt_row'] if is_alt else C['white'])
            cr.font = font(color='3B1A5A', size=9)
            cr.alignment = align()
            cr.border = thin_border()
            dv_acc.add(cr)
            # Año
            cy = ws.cell(row=excel_row, column=base+1, value=None)
            cy.fill = fill(C['alt_row'] if is_alt else C['white'])
            cy.font = font(color='3B1A5A', size=9)
            cy.alignment = align()
            cy.border = thin_border()
            cy.number_format = '0'
            dv_year.add(cy)
        # Otras (Realizada + Año + Descripción)
        ob = rest_start + (len(ACCIONES)-1)*2
        for off, fmt_str in enumerate([None, '0', None]):
            c = ws.cell(row=excel_row, column=ob+off, value='No' if off==0 else None)
            c.fill = fill(C['alt_row'] if is_alt else C['white'])
            c.font = font(color='3B1A5A', size=9)
            c.alignment = align(h='left' if off==2 else 'center')
            c.border = thin_border()
            if fmt_str: c.number_format = fmt_str
            if off == 0: dv_acc.add(c)
            if off == 1: dv_year.add(c)

    # ── Anchos de columna ─────────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(COL_TRAMO)].width = 7
    ws.column_dimensions[get_column_letter(COL_LARGO)].width = 9
    ws.column_dimensions[get_column_letter(COL_VUL)].width  = 13
    ws.column_dimensions[get_column_letter(COL_ACC)].width  = 12
    for i in range(len(NDVI_YEARS)):
        ws.column_dimensions[get_column_letter(ndvi_start+i)].width = 8
    ws.column_dimensions[get_column_letter(COL_TEND)].width  = 12
    ws.column_dimensions[get_column_letter(COL_SLOPE)].width = 10
    ws.column_dimensions[get_column_letter(COL_MKP)].width   = 9
    for i in range(len(INSTRUMENTOS)):
        ws.column_dimensions[get_column_letter(inst_start+i)].width = 16
    for i in range(len(ACCIONES)-1):
        ws.column_dimensions[get_column_letter(rest_start+i*2)].width   = 12
        ws.column_dimensions[get_column_letter(rest_start+i*2+1)].width = 7
    ob = rest_start + (len(ACCIONES)-1)*2
    ws.column_dimensions[get_column_letter(ob)].width   = 12
    ws.column_dimensions[get_column_letter(ob+1)].width = 7
    ws.column_dimensions[get_column_letter(ob+2)].width = 22

    # Alturas de fila
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 40
    ws.row_dimensions[3].height = 18
    for ridx in range(len(dept_df)):
        ws.row_dimensions[ridx+4].height = 16

    # Auto-filtro desde fila 3
    ws.auto_filter.ref = f'A3:{get_column_letter(TOTAL_COLS)}{len(dept_df)+3}'
    ws.sheet_view.zoomScale = 90

# ── Hoja RESUMEN ──────────────────────────────────────────────────────────────
ws_res = wb.create_sheet(title='Resumen', index=0)
ws_res.sheet_view.zoomScale = 90

set_header_cell(ws_res, 1, 1, 'KPI VEGETACIÓN COSTERA — RESUMEN POR DEPARTAMENTO',
                C['header_bg'], size=13, bold=True, wrap=False)
ws_res.merge_cells('A1:H1')
ws_res.row_dimensions[1].height = 30

hdrs = ['Departamento','Tramos','Vul. Alta','Vul. Media','Vul. Baja',
        'Mejora','Estable','Degradación']
for j, h in enumerate(hdrs, 1):
    set_header_cell(ws_res, 2, j, h, C['header_bg'], size=10)

for ri, dept_key in enumerate(DEPTS, 1):
    dept_df_s = df[df['d'] == dept_key]
    row = ri + 2
    is_alt = ri % 2 == 0
    bg = C['alt_row'] if is_alt else C['white']
    vals = [
        DEPT_NAMES[dept_key],
        len(dept_df_s),
        int((dept_df_s['v']==1).sum()),
        int((dept_df_s['v']==2).sum()),
        int((dept_df_s['v']==3).sum()),
        int((dept_df_s['tendencia']=='mejora').sum()),
        int((dept_df_s['tendencia']=='estable').sum()),
        int((dept_df_s['tendencia']=='degradacion').sum()),
    ]
    for j, val in enumerate(vals, 1):
        c = ws_res.cell(row=row, column=j, value=val)
        c.fill = fill(bg)
        c.font = font(bold=(j==1), size=10)
        c.alignment = align(h='left' if j==1 else 'center')
        c.border = thin_border()

# Totales
tot_row = len(DEPTS) + 3
set_header_cell(ws_res, tot_row, 1, 'TOTAL', C['header_bg'], size=10)
for j in range(2, 9):
    col_l = get_column_letter(j)
    c = ws_res.cell(row=tot_row, column=j,
                    value=f'=SUM({col_l}3:{col_l}{tot_row-1})')
    c.fill = fill(C['header_bg'])
    c.font = font(bold=True, color='EEE8DC', size=10)
    c.alignment = align()
    c.border = thin_border()

for j in range(1, 9):
    ws_res.column_dimensions[get_column_letter(j)].width = 18
ws_res.column_dimensions['A'].width = 18
ws_res.row_dimensions[1].height = 30

# ── Guardar ───────────────────────────────────────────────────────────────────
outpath = '/home/gaitapi/proyectos/kpi-ndvi-costera/data/KPI_Gestion_Costera.xlsx'
wb.save(outpath)
print(f'Guardado: {outpath}')
print(f'Hojas: {wb.sheetnames}')
